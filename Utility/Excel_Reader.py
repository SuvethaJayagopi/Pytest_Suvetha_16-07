import openpyxl

def get_data(path, sheet_name):
    try:
        print(f"Loading workbook from path: {path}")
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        final_list = []
        row = sheet.max_row
        col = sheet.max_column
        
        for r in range(2, row + 1):
            row_list = []
            for c in range(1, col + 1):
                row_list.append(sheet.cell(r, c).value)
            final_list.append(row_list)

        return final_list
    
    except Exception as e:
        print(f"Error loading workbook: {e}")
        raise
